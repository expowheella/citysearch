from openpyxl import load_workbook
import requests
from cities.settings import BASE_DIR



def test_1():
    file = BASE_DIR/'hundred_thousand_cities_full.xlsx'
    wb = load_workbook(file)
    # первый лист в книге
    ws = wb.active

    try:
        f = open(BASE_DIR/'result.txt', "r+")
        f.truncate(0)
        f.close()
    except Exception:
        pass

    with open(BASE_DIR/'result.txt', 'a', encoding="UTF-8") as f:
        for row in ws.iter_rows(min_row=88780):
            for cell in row:
                r = requests.get('http://127.0.0.1:8000/', data={'name': cell.value})
                print(f"{cell.value} -> {r.text}")
                f.write(cell.value + r.text + "\n")
    f.close

test_1()


def test_2():
    r = requests.get('http://127.0.0.1:8000/', data={'name': "Mazara Del Vallo"})
    print(f"{r.text}")

# test_2()